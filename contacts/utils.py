"""Utilitários de importação: normalização de telefone e leitura de planilhas."""
import csv
import io

import phonenumbers
from openpyxl import load_workbook


def normalize_phone(raw: str, region: str = "BR") -> str | None:
    """Valida e normaliza um telefone para o formato E.164.

    Retorna ``None`` quando o valor não é um telefone válido.
    """
    if not raw:
        return None
    raw = str(raw).strip()
    if not raw:
        return None
    try:
        parsed = phonenumbers.parse(raw, region)
    except phonenumbers.NumberParseException:
        return None
    if not phonenumbers.is_valid_number(parsed):
        return None
    return phonenumbers.format_number(parsed, phonenumbers.PhoneNumberFormat.E164)


def read_header_and_rows(filename: str, content: bytes) -> tuple[list[str], list[list[str]]]:
    """Lê um arquivo .csv ou .xlsx e retorna (cabeçalho, linhas) como texto.

    Levanta ``ValueError`` para formatos não suportados.
    """
    lower = filename.lower()
    if lower.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = [row for row in reader if any(cell.strip() for cell in row)]
        if not rows:
            return [], []
        header, *data_rows = rows
        return header, data_rows

    if lower.endswith(".xlsx"):
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = []
        for row in sheet.iter_rows(values_only=True):
            if any(cell not in (None, "") for cell in row):
                rows.append(["" if cell is None else str(cell) for cell in row])
        if not rows:
            return [], []
        header, *data_rows = rows
        return header, data_rows

    raise ValueError("Formato de arquivo não suportado. Envie um arquivo .csv ou .xlsx.")
